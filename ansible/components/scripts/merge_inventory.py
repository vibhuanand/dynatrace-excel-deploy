# ansible/components/scripts/merge_inventory.py
#!/usr/bin/env python3
import argparse, os, sys, json
from collections import OrderedDict
from zipfile import BadZipFile

# ---------------- deps ----------------
try:
    import yaml  # PyYAML
except Exception:
    print("ERROR: PyYAML is required (pip install pyyaml)", file=sys.stderr)
    sys.exit(2)

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # We'll fail later if an .xlsx/.xlsm is actually present

# ---------------- helpers ----------------
def _as_list(x):
    if x is None: return []
    if isinstance(x, (list, tuple)): return list(x)
    return [x]

def _is_nil(x) -> bool:
    return isinstance(x, str) and x.strip().lower() == "nil"

def _norm_header(h):
    return (h or "").strip().lower().replace(" ", "").replace("_", "")

def _to_bool(val):
    if val is None: return None
    s = str(val).strip().lower()
    if s in ("true","t","yes","y","1","on"):  return True
    if s in ("false","f","no","n","0","off"): return False
    return None

HEADER_MAP = {
    "host": {"host","hostname","host name"},
    "hostmatch": {"hostmatch","host_match","matchmode","hostmode"},
    "processname": {"process","processname","procname","name","process_name"},
    "enabled": {"enabled","toggle","onoff","active"},
    "matcherkey": {"matcherkey","key","matchkey","matcher","third","thirdarg","arg3"},
    "matchervalue": {"matchervalue","value","matchvalue","matcher_val","thirdvalue","thirdval","third_arg","arg"},
    "min": {"min","minimum","count","minprocs","minimumnumofprocesstobefound"},
    "os": {"os","platform"},
    "hostsuffix": {"hostsuffix","suffix","namesuffix"},
    "hostprefix": {"hostprefix","prefix","nameprefix"},
}

def canonicalize_headers(fieldnames):
    cmap = {}
    for f in fieldnames:
        nf = _norm_header(f)
        mapped = None
        for canon, alts in HEADER_MAP.items():
            if nf in alts:
                mapped = canon
                break
        cmap[f] = mapped or nf
    return cmap

def _split_os(val):
    if val is None or val == "": return []
    if _is_nil(val): return None  # explicit clear
    return [p.strip().upper() for p in str(val).split(",") if p and p.strip()]

def _is_lfs_pointer(path: str) -> bool:
    """Quick sniff for Git LFS pointer text file."""
    try:
        with open(path, "rb") as f:
            head = f.read(512)
        return b"git-lfs.github.com/spec" in head
    except Exception:
        return False

def _exit_lfs_pointer(path: str):
    print(
        f"ERROR: '{path}' looks like a Git LFS pointer, not a real Excel file.\n"
        f"Fix: enable LFS for your checkout and pull large files.\n"
        f"  - In Azure Pipelines YAML: add `lfs: true` to the `checkout: self` step\n"
        f"  - Or run on the agent: `git lfs install && git lfs fetch --all && git lfs pull`",
        file=sys.stderr
    )
    sys.exit(3)

# ---- metadata normalizer/merger (SSOT map → list[items]) ----
def normalize_metadata(md):
    """
    Accepts:
      - dict: {k:v}
      - list of 'k=v' strings
      - list of {'metadataKey':k,'metadataValue':v} dicts
    Returns list[{'metadataKey':..., 'metadataValue':...}]
    """
    items = []
    if not md:
        return items
    if isinstance(md, dict):
        for k, v in md.items():
            items.append({"metadataKey": str(k), "metadataValue": "" if v is None else str(v)})
    elif isinstance(md, list):
        for it in md:
            if isinstance(it, str) and "=" in it:
                k, v = it.split("=", 1)
                items.append({"metadataKey": k.strip(), "metadataValue": v.strip()})
            elif isinstance(it, dict) and "metadataKey" in it and "metadataValue" in it:
                items.append({
                    "metadataKey": str(it["metadataKey"]),
                    "metadataValue": "" if it.get("metadataValue") is None else str(it.get("metadataValue"))
                })
    return items

def merge_metadata_lists(a, b):
    """Merge two normalized metadata lists, de-duping by key (b overwrites a)."""
    out = {}
    for it in (a or []):
        out[str(it["metadataKey"])] = str(it.get("metadataValue", ""))
    for it in (b or []):
        out[str(it["metadataKey"])] = str(it.get("metadataValue", ""))
    return [{"metadataKey": k, "metadataValue": v} for k, v in out.items()]

def deep_merge_defaults(base, add):
    if not isinstance(base, dict): return add or {}
    if not isinstance(add, dict): return base or {}
    out = dict(base)
    for k, v in (add or {}).items():
        if k == "metadata":
            base_norm = normalize_metadata(out.get("metadata"))
            add_norm  = normalize_metadata(v)
            out["metadata"] = merge_metadata_lists(base_norm, add_norm)
        elif k == "os":
            base_os = set(_as_list(out.get("os")))
            add_os  = set(_as_list(v))
            out["os"] = sorted(base_os.union(add_os)) if add_os else (sorted(base_os) if base_os else _as_list(out.get("os")))
        else:
            out[k] = v
    return out

# ---------------- source parsers ----------------
def parse_json_file(path):
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {
        "nameSuffix": data.get("nameSuffix"),
        "namePrefix": data.get("namePrefix"),
        "defaults": data.get("defaults") or {},
        "hosts": data.get("hosts") or []
    }

def parse_xlsx_file(path, per_file_defaults):
    if load_workbook is None:
        print("ERROR: openpyxl is required for .xlsx/.xlsm (pip install openpyxl)", file=sys.stderr)
        sys.exit(2)

    if _is_lfs_pointer(path):
        _exit_lfs_pointer(path)

    try:
        wb = load_workbook(filename=path, data_only=True)
    except BadZipFile:
        if _is_lfs_pointer(path):
            _exit_lfs_pointer(path)
        print(
            f"ERROR: '{path}' is not a valid Excel zip (BadZipFile). "
            f"This often means the file is an unfinished Git LFS pointer.\n"
            f"Fix: enable LFS for checkout or run `git lfs pull` on the agent.",
            file=sys.stderr
        )
        sys.exit(3)

    ws = wb.active

    header_row = [(c.value if c.value is not None else "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = canonicalize_headers(header_row)

    # buckets hold mode per host + processes; **dedupe by host+process only**
    host_buckets = {}  # host -> {"hostMatch": <mode>, "processes": [], "_seen_proc": set()}
    os_accum = set()
    name_suffix = None
    name_prefix = None

    default_hmatch = (per_file_defaults.get("host_match") or "exact").strip().lower()
    if default_hmatch not in ("exact","contains","regex"):
        default_hmatch = "exact"

    priority = {"exact": 0, "contains": 1, "regex": 2}
    pfd_enabled = _to_bool(per_file_defaults.get("enabled"))  # <— NEW: per-file default

    for r in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        for idx, val in enumerate(r):
            key = header_map.get(header_row[idx]) or header_row[idx]
            row[key] = (str(val).strip() if val is not None else "")

        # Required fields
        host  = row.get("host") or ""
        pname = row.get("processname") or ""
        mval  = row.get("matchervalue") or ""
        if not host or not pname or not mval:
            continue

        # host bucket
        bucket = host_buckets.setdefault(host, {"hostMatch": default_hmatch, "processes": [], "_seen_proc": set()})

        # host match: row -> per-file default (keep the most permissive seen)
        hmatch = (row.get("hostmatch") or default_hmatch).strip().lower()
        if hmatch not in ("exact","contains","regex"):
            hmatch = default_hmatch
        if priority[hmatch] > priority.get(bucket["hostMatch"], 0):
            bucket["hostMatch"] = hmatch

        # suffix/prefix (first non-nil wins for the file)
        rsuf = row.get("hostsuffix", "")
        rpre = row.get("hostprefix", "")
        if name_suffix is None:
            if rsuf and not _is_nil(rsuf):
                name_suffix = rsuf
            elif "host_suffix" in per_file_defaults and not _is_nil(per_file_defaults.get("host_suffix")):
                name_suffix = per_file_defaults.get("host_suffix")
        if name_prefix is None:
            if rpre and not _is_nil(rpre):
                name_prefix = rpre
            elif "host_prefix" in per_file_defaults and not _is_nil(per_file_defaults.get("host_prefix")):
                name_prefix = per_file_defaults.get("host_prefix")

        # OS union for file defaults (nil disables contribution)
        ros = _split_os(row.get("os"))
        if ros is not None:
            os_accum.update(ros)

        # Build process (dedupe on host+process ONLY; ignore matcher/min differences)
        mkey = row.get("matcherkey") or per_file_defaults.get("matcher_key") or "cmd_contains"
        rmin = row.get("min")
        ren  = _to_bool(row.get("enabled"))  # <— NEW: row-level enabled

        sig = pname.lower()
        if sig in bucket["_seen_proc"]:
            continue
        bucket["_seen_proc"].add(sig)

        proc = OrderedDict()
        proc["name"] = pname
        proc["match"] = { mkey: mval }
        if rmin and not _is_nil(rmin):
            try:
                proc["min"] = int(str(rmin))
            except Exception:
                pass
        if ren is not None:
            proc["enabled"] = ren            # <— NEW

        bucket["processes"].append(proc)

    defaults = {}
    if "min" in per_file_defaults:
        defaults["min"] = per_file_defaults["min"]
    base_os = set(_as_list(per_file_defaults.get("os") or []))
    if base_os or os_accum:
        defaults["os"] = sorted(base_os.union(os_accum))
    if "metadata" in per_file_defaults:
        defaults["metadata"] = normalize_metadata(per_file_defaults["metadata"])
    if pfd_enabled is not None:
        defaults["enabled"] = pfd_enabled    # <— NEW

    hosts = [
        { "names": [h], "hostMatch": v["hostMatch"], "processes": v["processes"] }
        for h, v in host_buckets.items()
        if v.get("processes")
    ]

    return {
        "nameSuffix": name_suffix,
        "namePrefix": name_prefix,
        "defaults": defaults,
        "hosts": hosts
    }

# ---------------- global de-dupe (host+process ONLY) ----------------
def dedupe_combined_hosts(struct):
    """
    Collapse duplicates ACROSS ALL SOURCES strictly by:
      (host name, process name)
    """
    hosts = struct.get("hosts") or []
    merged = {}
    for hg in hosts:
        host_match = (hg.get("hostMatch") or "exact").strip().lower()
        names = hg.get("names") or []
        procs = hg.get("processes") or []
        for host in names:
            g = merged.setdefault(
                (host, host_match),
                {"host": host, "hostMatch": host_match, "processes": [], "_seen_proc": set()}
            )
            for p in procs:
                proc_name = (p.get("name") or "").strip()
                sig = proc_name.lower()
                if sig in g["_seen_proc"]:
                    continue
                g["_seen_proc"].add(sig)
                g["processes"].append(p)

    struct["hosts"] = [
        {"names": [v["host"]], "hostMatch": v["hostMatch"], "processes": v["processes"]}
        for v in merged.values()
        if v["processes"]
    ]
    return struct

# ---------------- main ----------------
def main():
    ap = argparse.ArgumentParser(description="Merge inventory sources (JSON + Excel) into one canonical JSON")
    ap.add_argument("--ssot", required=True, help="Path to env YAML (SSOT)")
    ap.add_argument("--out", required=True, help="Output path for merged canonical JSON")
    args = ap.parse_args()

    if not os.path.isfile(args.ssot):
        print(f"ERROR: SSOT not found: {args.ssot}", file=sys.stderr)
        sys.exit(2)

    cfg = yaml.safe_load(open(args.ssot, "r", encoding="utf-8")) or {}
    res = (cfg.get("resources") or {}).get("process_availability")

    # ---------- SSOT-level defaults ----------
    ssot_defaults = {}
    if isinstance(res, dict):
        ssot_defaults = res.get("defaults") or {}
    if "metadata" in ssot_defaults:
        ssot_defaults = dict(ssot_defaults)
        ssot_defaults["metadata"] = normalize_metadata(ssot_defaults["metadata"])

    sources_raw = []
    if res and isinstance(res.get("sources"), list):
        sources_raw = res["sources"]
    else:
        legacy = (cfg.get("resources") or {}).get("process_monitoring_inventory")
        if legacy:
            sources_raw = [{"path": legacy, "defaults": {}}]

    sources = []
    for s in sources_raw:
        if isinstance(s, str):
            sources.append({"path": s, "defaults": {}})
        elif isinstance(s, dict) and "path" in s:
            s.setdefault("defaults", {})
            sources.append(s)

    if not sources:
        print("ERROR: No sources defined in resources.process_availability.sources", file=sys.stderr)
        sys.exit(2)

    combined = {
        "version": 1,
        "nameSuffix": None,
        "namePrefix": None,
        "defaults": ssot_defaults.copy(),   # start from SSOT defaults
        "hosts": []
    }

    first_suffix = None
    first_prefix = None

    ssot_dir = os.path.dirname(os.path.abspath(args.ssot))

    for entry in sources:
        rel = entry["path"]
        per_file_defaults = entry.get("defaults") or {}
        p = rel if os.path.isabs(rel) else os.path.abspath(os.path.join(ssot_dir, rel))
        if not os.path.isfile(p):
            print(f"WARNING: inventory source missing: {p}", file=sys.stderr)
            continue

        ext = os.path.splitext(p)[1].lower()

        if ext == ".json":
            part = parse_json_file(p)
            # normalize metadata on both sides then merge defaults
            if "metadata" in part.get("defaults", {}):
                part["defaults"]["metadata"] = normalize_metadata(part["defaults"]["metadata"])
            pfd = dict(per_file_defaults)
            if "metadata" in pfd:
                pfd["metadata"] = normalize_metadata(pfd["metadata"])
            fdefs = deep_merge_defaults(part.get("defaults") or {}, pfd)
            part["defaults"] = fdefs

            # suffix/prefix overrides
            if "host_suffix" in per_file_defaults:
                part["nameSuffix"] = None if _is_nil(per_file_defaults["host_suffix"]) else per_file_defaults["host_suffix"]
            if "host_prefix" in per_file_defaults:
                part["namePrefix"] = None if _is_nil(per_file_defaults["host_prefix"]) else per_file_defaults["host_prefix"]

            # ensure hostMatch on JSON groups from per-file default
            default_hmatch = (per_file_defaults.get("host_match") or "exact").strip().lower()
            if default_hmatch not in ("exact","contains","regex"):
                default_hmatch = "exact"
            for hg in (part.get("hosts") or []):
                if "hostMatch" not in hg or not hg.get("hostMatch"):
                    hg["hostMatch"] = default_hmatch

        elif ext in (".xlsx", ".xlsm"):
            try:
                part = parse_xlsx_file(p, per_file_defaults)
            except SystemExit as e:
                raise
            except Exception as e:
                msg = (
                    f"ERROR: Failed reading Excel source '{p}': {e}\n"
                    f"HINT: If this file is stored with Git LFS, make sure it was downloaded "
                    f"(use `checkout: self` with `lfs: true` in Pipelines, or run `git lfs pull`)."
                )
                print(msg, file=sys.stderr)
                sys.exit(3)

        else:
            print(f"WARNING: unsupported source type (only .json/.xlsx/.xlsm): {p}", file=sys.stderr)
            continue

        if first_suffix is None and part.get("nameSuffix") is not None:
            first_suffix = part["nameSuffix"]
        if first_prefix is None and part.get("namePrefix") is not None:
            first_prefix = part["namePrefix"]

        # merge source defaults into combined defaults (SSOT base -> source layer)
        combined["defaults"] = deep_merge_defaults(combined.get("defaults") or {}, part.get("defaults") or {})

        combined["hosts"].extend(part.get("hosts") or [])

    combined["nameSuffix"] = first_suffix
    combined["namePrefix"] = first_prefix

    # GLOBAL DEDUPE per your rule (host+process only)
    combined = dedupe_combined_hosts(combined)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(combined, f, indent=2, ensure_ascii=False)

    print(f"Wrote merged inventory: {args.out}")
    print(f" - namePrefix: {combined.get('namePrefix')}")
    print(f" - nameSuffix: {combined.get('nameSuffix')}")
    defs = combined.get("defaults", {})
    print(f" - defaults.min: {defs.get('min')}")
    print(f" - defaults.os:  {defs.get('os')}")
    if "enabled" in defs:
        print(f" - defaults.enabled: {defs.get('enabled')}")  # <— helpful log
    md = defs.get("metadata")
    if md:
        keys = [it.get("metadataKey") for it in md if isinstance(it, dict)]
        print(f" - defaults.metadata keys: {keys}")
    print(f" - host groups:  {len(combined.get('hosts') or [])}")

if __name__ == "__main__":
    main()
