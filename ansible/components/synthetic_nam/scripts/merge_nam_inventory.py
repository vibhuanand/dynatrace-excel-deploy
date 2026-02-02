#!/usr/bin/env python3
import os, sys, json, argparse
try:
    import yaml
except Exception:
    print("PyYAML required", file=sys.stderr); sys.exit(2)

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # will error if .xlsx is used

def _split_csv(s):
    if not s: return []
    return [p.strip() for p in str(s).split(",") if p and p.strip()]

def _read_json(p):
    with open(p, "r", encoding="utf-8") as f: return json.load(f)

def _read_xlsx(p):
    if load_workbook is None:
        print("openpyxl required for .xlsx", file=sys.stderr); sys.exit(2)
    wb = load_workbook(p, data_only=True); ws = wb.active
    head = [c.value if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = { (h or "").strip().lower(): i for i, h in enumerate(head) }
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        g = lambda k: r[idx[k]] if k in idx else None
        out.append({
            "name": g("name"),
            "targets": _split_csv(g("targets") or ""),
            "locations": _split_csv(g("locations") or ""),
            "frequency_min": int(g("frequency_min") or 5),
            "icmp": {
                "packets": int(g("icmp_packets") or 5),
                "timeout": (g("icmp_timeout") or "PT1S"),
                "ip_version": str(g("icmp_ip_version") or "4"),
            },
            "enabled": str(g("enabled") or "true").strip().lower() in ("true","1","yes","y")
        })
    return [x for x in out if x.get("name") and x.get("targets")]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ssot", required=True)   # envs/...yaml
    ap.add_argument("--out",  required=True)   # output canonical json
    args = ap.parse_args()

    cfg = yaml.safe_load(open(args.ssot, "r", encoding="utf-8")) or {}
    res = ((cfg.get("resources") or {}).get("network_availability") or {})
    defaults = res.get("defaults") or {}
    sources  = res.get("sources") or []

    items = []
    base_locs = defaults.get("locations") or []
    base_freq = int(defaults.get("frequency_min") or 5)
    base_icmp = (defaults.get("icmp") or {})
    base_cons = (defaults.get("constraints") or {})
    outage    = (defaults.get("outage_handling") or {})

    for s in sources:
        path = s["path"] if isinstance(s, dict) else s
        p = path if os.path.isabs(path) else os.path.join(os.path.dirname(args.ssot), path)
        if not os.path.isfile(p): 
            print(f"WARNING: missing source {p}", file=sys.stderr); 
            continue
        ext = os.path.splitext(p)[1].lower()
        rows = _read_json(p) if ext == ".json" else _read_xlsx(p) if ext in (".xlsx",".xlsm") else []
        for r in rows:
            itm = {
                "name": r.get("name"),
                "enabled": bool(r.get("enabled", True)),
                "frequencyMin": int(r.get("frequency_min") or base_freq),
                "locations": r.get("locations") or base_locs,
                "icmp": {
                    "packets": int((r.get("icmp") or {}).get("packets", base_icmp.get("packets", 5))),
                    "timeout": (r.get("icmp") or {}).get("timeout", base_icmp.get("timeout", "PT1S")),
                    "ip_version": str((r.get("icmp") or {}).get("ip_version", base_icmp.get("ip_version", "4")))
                },
                "constraints": {
                    "step_success_rate": r.get("constraints", {}).get("step_success_rate", base_cons.get("step_success_rate", {"operator":">=","value":"95"})),
                    "per_target_success": r.get("constraints", {}).get("per_target_success", base_cons.get("per_target_success", {"operator":"=","value":"100"}))
                },
                "outage_handling": outage,
                "targets": r.get("targets") or []
            }
            if itm["name"] and itm["targets"]:
                items.append(itm)

    os.makedirs(os.path.dirname(args.out), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump({"monitors": items}, f, indent=2, ensure_ascii=False)
    print(f"Wrote NAM canonical: {args.out}  (count={len(items)})")

if __name__ == "__main__":
    main()
